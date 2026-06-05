#!/usr/bin/env python3
import sys
sys.path.insert(0, 'C:\\Users\\sbirnbaum\\AppData\\Roaming\\Python\\Python312\\site-packages')

try:
    import openpyxl
    wb = openpyxl.load_workbook('Feeder Schedule - Wire Size Chart 2026-06-01).xlsx')
    ws = wb.active

    print("Sheet name:", ws.title)
    print("Max row:", ws.max_row)
    print("\nFirst 20 rows:")

    for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
        print(row)

except ImportError as e:
    print(f"Import error: {e}")
    print("openpyxl not found, trying alternate method...")

    try:
        import pandas as pd
        df = pd.read_excel('Feeder Schedule - Wire Size Chart 2026-06-01).xlsx')
        print(df.head(60).to_string())
    except Exception as e2:
        print(f"Pandas error: {e2}")
