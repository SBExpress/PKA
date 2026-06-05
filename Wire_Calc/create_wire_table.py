import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

nec_310_16 = {
    '14 AWG': {'copper': 20, 'aluminum': 0},
    '12 AWG': {'copper': 25, 'aluminum': 20},
    '10 AWG': {'copper': 35, 'aluminum': 30},
    '8 AWG': {'copper': 50, 'aluminum': 40},
    '6 AWG': {'copper': 65, 'aluminum': 50},
    '4 AWG': {'copper': 85, 'aluminum': 65},
    '3 AWG': {'copper': 100, 'aluminum': 75},
    '2 AWG': {'copper': 115, 'aluminum': 90},
    '1 AWG': {'copper': 130, 'aluminum': 100},
    '1/0 AWG': {'copper': 150, 'aluminum': 120},
    '2/0 AWG': {'copper': 175, 'aluminum': 135},
    '3/0 AWG': {'copper': 200, 'aluminum': 155},
    '4/0 AWG': {'copper': 230, 'aluminum': 180},
    '250 kcmil': {'copper': 255, 'aluminum': 205},
    '300 kcmil': {'copper': 285, 'aluminum': 230},
    '350 kcmil': {'copper': 310, 'aluminum': 250},
    '400 kcmil': {'copper': 335, 'aluminum': 270},
    '500 kcmil': {'copper': 380, 'aluminum': 310},
    '600 kcmil': {'copper': 420, 'aluminum': 340},
    '700 kcmil': {'copper': 460, 'aluminum': 375},
    '750 kcmil': {'copper': 475, 'aluminum': 385},
    '800 kcmil': {'copper': 490, 'aluminum': 395},
    '900 kcmil': {'copper': 520, 'aluminum': 425},
    '1000 kcmil': {'copper': 545, 'aluminum': 445},
    '1250 kcmil': {'copper': 590, 'aluminum': 485},
    '1500 kcmil': {'copper': 625, 'aluminum': 520},
    '1750 kcmil': {'copper': 650, 'aluminum': 545},
    '2000 kcmil': {'copper': 665, 'aluminum': 560},
}

resistance = {
    '14 AWG': {'copper': 3.07, 'aluminum': 5.06},
    '12 AWG': {'copper': 1.93, 'aluminum': 3.18},
    '10 AWG': {'copper': 1.21, 'aluminum': 2.0},
    '8 AWG': {'copper': 0.764, 'aluminum': 1.26},
    '6 AWG': {'copper': 0.491, 'aluminum': 0.808},
    '4 AWG': {'copper': 0.308, 'aluminum': 0.508},
    '3 AWG': {'copper': 0.245, 'aluminum': 0.403},
    '2 AWG': {'copper': 0.194, 'aluminum': 0.319},
    '1 AWG': {'copper': 0.154, 'aluminum': 0.253},
    '1/0 AWG': {'copper': 0.122, 'aluminum': 0.201},
    '2/0 AWG': {'copper': 0.0967, 'aluminum': 0.159},
    '3/0 AWG': {'copper': 0.0766, 'aluminum': 0.126},
    '4/0 AWG': {'copper': 0.0608, 'aluminum': 0.1},
    '250 kcmil': {'copper': 0.0515, 'aluminum': 0.0847},
    '300 kcmil': {'copper': 0.0429, 'aluminum': 0.0707},
    '350 kcmil': {'copper': 0.0367, 'aluminum': 0.0605},
    '400 kcmil': {'copper': 0.0321, 'aluminum': 0.0529},
    '500 kcmil': {'copper': 0.0258, 'aluminum': 0.0424},
    '600 kcmil': {'copper': 0.0214, 'aluminum': 0.0353},
    '700 kcmil': {'copper': 0.0184, 'aluminum': 0.0303},
    '750 kcmil': {'copper': 0.0171, 'aluminum': 0.0282},
    '800 kcmil': {'copper': 0.0161, 'aluminum': 0.0265},
    '900 kcmil': {'copper': 0.0143, 'aluminum': 0.0235},
    '1000 kcmil': {'copper': 0.0129, 'aluminum': 0.0212},
    '1250 kcmil': {'copper': 0.01033, 'aluminum': 0.01700},
    '1500 kcmil': {'copper': 0.00860, 'aluminum': 0.01410},
    '1750 kcmil': {'copper': 0.00736, 'aluminum': 0.01210},
    '2000 kcmil': {'copper': 0.00645, 'aluminum': 0.01060},
}

conduit_sizes = {
    '14 AWG': '1/2"', '12 AWG': '1/2"', '10 AWG': '1/2"', '8 AWG': '1/2"',
    '6 AWG': '1/2"', '4 AWG': '3/4"', '3 AWG': '3/4"', '2 AWG': '3/4"',
    '1 AWG': '1"', '1/0 AWG': '1"', '2/0 AWG': '1-1/4"', '3/0 AWG': '1-1/4"',
    '4/0 AWG': '1-1/2"', '250 kcmil': '1-1/2"', '300 kcmil': '2"',
    '350 kcmil': '2"', '400 kcmil': '2"', '500 kcmil': '2-1/2"',
    '600 kcmil': '2-1/2"', '700 kcmil': '3"', '750 kcmil': '3"',
    '800 kcmil': '3"', '900 kcmil': '3-1/2"', '1000 kcmil': '3-1/2"',
    '1250 kcmil': '4"', '1500 kcmil': '4"', '1750 kcmil': '4-1/2"', '2000 kcmil': '5"',
}

wire_sizes_ordered = [
    '14 AWG', '12 AWG', '10 AWG', '8 AWG', '6 AWG', '4 AWG', '3 AWG', '2 AWG', '1 AWG',
    '1/0 AWG', '2/0 AWG', '3/0 AWG', '4/0 AWG',
    '250 kcmil', '300 kcmil', '350 kcmil', '400 kcmil', '500 kcmil', '600 kcmil',
    '700 kcmil', '750 kcmil', '800 kcmil', '900 kcmil', '1000 kcmil',
    '1250 kcmil', '1500 kcmil', '1750 kcmil', '2000 kcmil',
]

def find_wire_size(amperage, conductor_type):
    for wire in wire_sizes_ordered:
        ampacity = nec_310_16.get(wire, {}).get(conductor_type, 0)
        if ampacity >= amperage:
            return wire
    return '2000 kcmil'

def calculate_max_length(amperage, wire_size, voltage, conductor_type):
    r = resistance.get(wire_size, {}).get(conductor_type, 0.1)
    if r == 0:
        return 0
    max_vd = voltage * 0.03
    sqrt3 = 1.732
    max_length = (max_vd * 1000) / (sqrt3 * r * amperage)
    return max(10, int(max_length / 10) * 10)

wb = Workbook()
ws = wb.active
ws.title = 'Wire Sizing'

header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
voltage_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
voltage_font = Font(bold=True, size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.merge_cells('A1:I1')
title = ws['A1']
title.value = 'NEC 310.16 - 3-Phase Wire Sizing Guide (75°C, 3% Voltage Drop, 30°C Ambient)'
title.font = Font(bold=True, size=14, color='FFFFFF')
title.fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
title.alignment = center_align

row = 3
ws.cell(row, 1).value = 'Amperage'
ws.cell(row, 1).font = header_font
ws.cell(row, 1).fill = header_fill
ws.cell(row, 1).alignment = center_align
ws.cell(row, 1).border = border

ws.merge_cells('B3:E3')
cell_208 = ws['B3']
cell_208.value = '208V 3-Phase'
cell_208.fill = voltage_fill
cell_208.font = voltage_font
cell_208.alignment = center_align
cell_208.border = border

ws.merge_cells('F3:I3')
cell_480 = ws['F3']
cell_480.value = '480V 3-Phase'
cell_480.fill = voltage_fill
cell_480.font = voltage_font
cell_480.alignment = center_align
cell_480.border = border

row = 4
for i, header in enumerate(['Amperage', 'Cu Wire', 'Al Wire', 'Conduit', 'Max Len (ft)', 'Cu Wire', 'Al Wire', 'Conduit', 'Max Len (ft)'], 1):
    cell = ws.cell(row, i)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

ws.column_dimensions['A'].width = 12
for i in range(2, 10):
    ws.column_dimensions[get_column_letter(i)].width = 13

data_row = 5
for amp in range(20, 3010, 10):
    ws.cell(data_row, 1).value = amp
    ws.cell(data_row, 1).alignment = center_align
    ws.cell(data_row, 1).border = border

    cu_208 = find_wire_size(amp, 'copper')
    al_208 = find_wire_size(amp, 'aluminum')
    conduit_208 = conduit_sizes.get(cu_208, '4"')
    max_len_208 = calculate_max_length(amp, cu_208, 208, 'copper')

    ws.cell(data_row, 2).value = cu_208
    ws.cell(data_row, 2).alignment = center_align
    ws.cell(data_row, 2).border = border

    ws.cell(data_row, 3).value = al_208
    ws.cell(data_row, 3).alignment = center_align
    ws.cell(data_row, 3).border = border

    ws.cell(data_row, 4).value = conduit_208
    ws.cell(data_row, 4).alignment = center_align
    ws.cell(data_row, 4).border = border

    ws.cell(data_row, 5).value = max_len_208
    ws.cell(data_row, 5).alignment = center_align
    ws.cell(data_row, 5).border = border

    cu_480 = find_wire_size(amp, 'copper')
    al_480 = find_wire_size(amp, 'aluminum')
    conduit_480 = conduit_sizes.get(cu_480, '4"')
    max_len_480 = calculate_max_length(amp, cu_480, 480, 'copper')

    ws.cell(data_row, 6).value = cu_480
    ws.cell(data_row, 6).alignment = center_align
    ws.cell(data_row, 6).border = border

    ws.cell(data_row, 7).value = al_480
    ws.cell(data_row, 7).alignment = center_align
    ws.cell(data_row, 7).border = border

    ws.cell(data_row, 8).value = conduit_480
    ws.cell(data_row, 8).alignment = center_align
    ws.cell(data_row, 8).border = border

    ws.cell(data_row, 9).value = max_len_480
    ws.cell(data_row, 9).alignment = center_align
    ws.cell(data_row, 9).border = border

    data_row += 1

footer_row = data_row + 2
ws.merge_cells(f'A{footer_row}:I{footer_row}')
notes_cell = ws[f'A{footer_row}']
notes_cell.value = 'Based on NEC 2023 Table 310.16 | 75°C insulation rating | 3% voltage drop limit | 30°C ambient | 3-phase balanced loads'
notes_cell.font = Font(italic=True, size=9)
notes_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[footer_row].height = 20

output_path = 'E:\\Shaya\\Claude AI\\PKA\\Wire_Sizing_Table.xlsx'
wb.save(output_path)
print(f'Created: {output_path}')
